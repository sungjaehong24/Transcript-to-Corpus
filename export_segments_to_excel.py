"""
Export Word (.docx) qualitative coding to Excel.

**Default:** one row per unique **quote** within a file. Several comments on the same
anchored text merge into that row; the **code** column lists labels joined with **'; '**
(in a stable project order). Highlight colours collapse near-synonyms (e.g. gray vs
dark gray; green vs bright green). All grey-shade highlights rely on comment wording
when appraisal codes apply, otherwise **Information Source**.

Mapping matches the project’s “List of codes” table: highlight colour (+ comment when
needed). Highlights **without** comments are exported when they fall outside any comment
anchor in the document body.

**Legacy:** `--coder-matrix` restores the older layout (one column per Word comment author).

Requires: openpyxl. Reads OOXML inside .docx (zip files). Offline; no APIs.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

# Word stores slightly different highlight names across versions/themes; collapse families.
_GRAY_HIGHLIGHTS = frozenset(
    {"lightgray", "lightgrey", "gray", "grey", "darkgray", "darkgrey"}
)
_GREEN_HIGHLIGHTS = frozenset({"green", "brightgreen", "darkgreen"})
_TEAL_HIGHLIGHTS = frozenset({"darkcyan", "teal"})


def _w(tag: str) -> str:
    return f"{{{W_NS}}}{tag}"


def _local(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag


def _attrib_val(attrs: dict) -> Optional[str]:
    for key, val in attrs.items():
        if key.endswith("}val") or key == "val":
            return val
    return None


def _attr_id(elem: ET.Element) -> Optional[str]:
    for key, val in elem.attrib.items():
        if key.endswith("}id") or key == "id":
            return val
    return None


def _collect_w_t_text(node: ET.Element) -> str:
    parts: List[str] = []
    for el in node.iter():
        if _local(el.tag) == "t":
            if el.text:
                parts.append(el.text)
            if el.tail:
                parts.append(el.tail)
    return "".join(parts)


def _parse_comments_xml(zf: zipfile.ZipFile) -> Dict[str, Tuple[str, str]]:
    try:
        data = zf.read("word/comments.xml")
    except KeyError:
        return {}
    root = ET.fromstring(data)
    out: Dict[str, Tuple[str, str]] = {}
    for el in root.iter():
        if _local(el.tag) != "comment":
            continue
        cid = _attr_id(el)
        if cid is None:
            continue
        author = el.get(_w("author")) or el.get("author") or ""
        body = _collect_w_t_text(el)
        body = re.sub(r"\s+", " ", body).strip()
        out[cid] = (author, body)
    return out


def _normalize_segment_text(s: str) -> str:
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _format_transcript_layout(s: str) -> str:
    if not s:
        return s
    return re.sub(
        r"((?:P02|Interviewer)\s+\d{1,2}:\d{2})\s*([^\n])",
        r"\1\n\2",
        s,
    )


def _flatten_doc(elem: ET.Element) -> List[ET.Element]:
    return list(elem.iter())


def _segment_anchor_indices(
    doc_root: ET.Element, comment_id: str
) -> Tuple[Optional[int], Optional[int]]:
    start_tag = _w("commentRangeStart")
    end_tag = _w("commentRangeEnd")
    flat = _flatten_doc(doc_root)
    si = ei = None
    for idx, el in enumerate(flat):
        if el.tag == start_tag and _attr_id(el) == comment_id:
            si = idx
        if el.tag == end_tag and _attr_id(el) == comment_id:
            ei = idx
            break
    return si, ei


def _all_comment_anchor_intervals(
    doc_root: ET.Element, comment_ids: Iterable[str]
) -> List[Tuple[int, int]]:
    spans: List[Tuple[int, int]] = []
    for cid in comment_ids:
        si, ei = _segment_anchor_indices(doc_root, str(cid))
        if si is None or ei is None or ei <= si:
            continue
        spans.append((si, ei))
    return spans


def _flat_index_inside_any_comment(
    idx: int, intervals: Sequence[Tuple[int, int]]
) -> bool:
    return any(si < idx < ei for si, ei in intervals)


def _segments_highlighted_outside_comments(
    doc_root: ET.Element,
    intervals: Sequence[Tuple[int, int]],
) -> List[Tuple[int, List[ET.Element]]]:
    """
    Word comment 앵커 *밖*에서만 이어지는 하이라이트 run 묶음마다 (flat시작인덱스, runs).
    """
    flat = _flatten_doc(doc_root)
    grouped: List[Tuple[int, List[ET.Element]]] = []
    buf: List[ET.Element] = []
    buf_fam: Optional[str] = None
    start_idx: Optional[int] = None

    def flush() -> None:
        nonlocal buf, buf_fam, start_idx
        if buf and start_idx is not None:
            grouped.append((start_idx, buf[:]))
        buf = []
        buf_fam = None
        start_idx = None

    for idx, el in enumerate(flat):
        if el.tag != _w("r"):
            continue
        if _flat_index_inside_any_comment(idx, intervals):
            flush()
            continue
        raw_h = _highlight_on_run(el)
        if not raw_h:
            flush()
            continue
        fam = _highlight_family_bucket(raw_h)
        if not buf:
            buf = [el]
            buf_fam = fam
            start_idx = idx
        elif fam == buf_fam:
            buf.append(el)
        else:
            flush()
            buf = [el]
            buf_fam = fam
            start_idx = idx
    flush()
    return grouped


def _rows_from_highlight_only_segments(
    hl_segments: Sequence[Tuple[int, List[ET.Element]]],
) -> List[dict]:
    rows: List[dict] = []
    for doc_pos, ordered_runs in hl_segments:
        chunks = _merged_highlight_chunks(ordered_runs)
        hl_dom = _dominant_highlight(chunks)
        quote = _quote_for_highlight(chunks, hl_dom) if hl_dom else ""
        quote_norm = _format_transcript_layout(_normalize_segment_text(quote))
        if not quote_norm.strip():
            continue
        code, warns = _code_from_highlight_and_comment(hl_dom, "")
        rows.append(
            {
                "comment_id": f"hl-{doc_pos}",
                "quote": quote_norm,
                "code": code.strip(),
                "raw_comment": "",
                "author": "",
                "_warns": warns,
                "_doc_order": doc_pos,
            },
        )
    return rows


def _runs_between_anchor_indices(
    doc_root: ET.Element, start_idx: int, end_idx: int
) -> List[ET.Element]:
    flat = _flatten_doc(doc_root)
    return [
        el
        for idx, el in enumerate(flat)
        if start_idx < idx < end_idx and el.tag == _w("r")
    ]


def _segment_text_for_comment(doc_root: ET.Element, comment_id: str) -> str:
    si, ei = _segment_anchor_indices(doc_root, comment_id)
    if si is None or ei is None or ei <= si:
        return ""
    flat = _flatten_doc(doc_root)
    parts: List[str] = []
    for idx in range(si + 1, ei):
        el = flat[idx]
        if _local(el.tag) == "t":
            if el.text:
                parts.append(el.text)
            if el.tail:
                parts.append(el.tail)
    return "".join(parts)


def _highlight_on_run(run: ET.Element) -> Optional[str]:
    for child in run:
        if _local(child.tag) != "rPr":
            continue
        for rp in child:
            if _local(rp.tag) != "highlight":
                continue
            v = _attrib_val(rp.attrib)
            if not v:
                continue
            vlow = str(v).strip().lower()
            return None if vlow == "none" else vlow
        break
    return None


def _collect_run_plain_text(run: ET.Element) -> str:
    return "".join(run.itertext())


def _norm_hi_text(s: str) -> str:
    s = s.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", s).strip()


def _highlight_family_bucket(raw: Optional[str]) -> str:
    """
    OOXML highlight token → semantic bucket ('gray', 'green', 'teal', …).
    Tokens not listed pass through lowercase (yellow, cyan, …).
    """
    if not raw:
        return ""
    h = str(raw).strip().lower()
    if h in _GRAY_HIGHLIGHTS:
        return "gray"
    if h in _GREEN_HIGHLIGHTS:
        return "green"
    if h in _TEAL_HIGHLIGHTS:
        return "teal"
    return h


_CODE_SORT_PRIORITY: Dict[str, int] = {
    name: idx
    for idx, name in enumerate(
        [
            "IBM",
            "Belonging",
            "Motivation",
            "Source Appraisal",
            "Information Appraisal",
            "Information Source",
            "Cues",
            "Event Appraisal",
            "Facilitators",
            "Barriers",
            "Other",
        ],
    )
}


def _canonical_sort_code_labels(labels: List[str]) -> List[str]:
    dedup = list(dict.fromkeys(labels))
    return sorted(dedup, key=lambda x: (_CODE_SORT_PRIORITY.get(x, 1000), x.lower()))


def _merged_highlight_chunks(
    ordered_runs: List[ET.Element],
) -> List[Tuple[str, str]]:
    """Contiguous runs sharing the same highlight *family* (gray/green/teal…) → chunk."""
    out: List[Tuple[str, str]] = []
    i = 0
    n = len(ordered_runs)
    while i < n:
        r = ordered_runs[i]
        hv = _highlight_on_run(r)
        if not hv:
            i += 1
            continue
        hv = hv.lower()
        fam = _highlight_family_bucket(hv)
        pieces = [_norm_hi_text(_collect_run_plain_text(r))]
        i += 1
        while i < n:
            r2 = ordered_runs[i]
            hv2 = _highlight_on_run(r2)
            if hv2 and _highlight_family_bucket(hv2) == fam:
                pieces.append(_norm_hi_text(_collect_run_plain_text(r2)))
                i += 1
                continue
            if not _norm_hi_text(_collect_run_plain_text(r2)):
                i += 1
                continue
            break
        joined = _norm_hi_text("".join(pieces))
        if joined:
            out.append((hv, joined))
    return out


def _dominant_highlight(chunks: List[Tuple[str, str]]) -> Optional[str]:
    """Winner by highlighted character count; merges grey/green families when chunking."""
    if not chunks:
        return None
    totals: Dict[str, Tuple[int, str]] = {}
    for c, tx in chunks:
        fam = _highlight_family_bucket(c)
        key = fam if fam else str(c).lower()
        ln = len(tx)
        if key not in totals:
            totals[key] = (ln, str(c))
        else:
            prev_len, prev_s = totals[key]
            totals[key] = (prev_len + ln, prev_s)
    win = max(totals.keys(), key=lambda k: totals[k][0])
    return totals[win][1]


def _quote_for_highlight(chunks: List[Tuple[str, str]], dominant_raw: str) -> str:
    if not dominant_raw:
        return ""
    df = _highlight_family_bucket(dominant_raw)
    parts: List[str] = []
    for hc, t in chunks:
        if df and _highlight_family_bucket(hc) == df:
            parts.append(t)
        elif not df and str(hc).lower() == str(dominant_raw).lower():
            parts.append(t)
    return _norm_hi_text(" ".join(parts))


def _codes_from_comment_only(clo: str) -> List[str]:
    """Highlight missing; recover IBM/Belonging or appraisal phrases from comment text."""
    codes: List[str] = []
    if "information appraisal" in clo:
        codes.append("Information Appraisal")
    if "source appraisal" in clo:
        codes.append("Source Appraisal")
    if codes:
        return _canonical_sort_code_labels(codes)
    g: List[str] = []
    if "ibm" in clo:
        g.append("IBM")
    if "belonging" in clo:
        g.append("Belonging")
    if g:
        return _canonical_sort_code_labels(g)
    return []


def _code_from_highlight_and_comment(
    dominant_hl: Optional[str],
    comment_body: str,
) -> Tuple[str, List[str]]:
    """
    Returns (code string, possibly multiple labels joined with '; ', stderr warnings).
    Gray / green families tolerate Word naming drift; multi-label when comment supports it.
    """
    warn: List[str] = []
    clo = comment_body.strip().lower()
    bucket = _highlight_family_bucket(dominant_hl)
    codes: List[str] = []

    if not bucket:
        codes = _codes_from_comment_only(clo)
        if not codes:
            return "", warn
        return "; ".join(_canonical_sort_code_labels(codes)), warn

    if bucket == "gray":
        if "information appraisal" in clo:
            codes.append("Information Appraisal")
        if "source appraisal" in clo:
            codes.append("Source Appraisal")
        if codes:
            return "; ".join(_canonical_sort_code_labels(codes)), warn
        return "Information Source", warn

    if bucket == "green":
        if "ibm" in clo:
            codes.append("IBM")
        if "belonging" in clo:
            codes.append("Belonging")
        if codes:
            return "; ".join(_canonical_sort_code_labels(codes)), warn
        return "Motivation", warn

    if bucket == "teal":
        return "Event Appraisal", warn

    if bucket == "yellow":
        return "Cues", warn

    if bucket == "cyan":
        return "Facilitators", warn

    if bucket == "red":
        return "Barriers", warn

    if bucket == "magenta":
        return "Other", warn

    warn.append(
        f"[warn] Unrecognized highlight colour for automatic mapping: {dominant_hl!r}",
    )
    return "", warn


def _cid_sort_tuple(cid: str) -> Tuple[int | str, str]:
    cid = str(cid)
    return (int(cid), cid) if cid.isdigit() else (cid, cid)


def _merge_highlight_rows_same_quote(rows: List[dict]) -> List[dict]:
    """Same quote + multiple Word comments → one row, codes '; '-joined."""
    buckets: Dict[str, List[dict]] = defaultdict(list)
    first_seen: Dict[str, int] = {}
    for pos, row in enumerate(rows):
        q = row["quote"]
        buckets[q].append(row)
        if q not in first_seen:
            first_seen[q] = pos
    out: List[dict] = []
    for quote in sorted(buckets.keys(), key=lambda qq: first_seen[qq]):
        grp = buckets[quote]
        grp.sort(key=lambda r: _cid_sort_tuple(str(r.get("comment_id", ""))))

        uniq: Dict[str, None] = {}
        merged_warns: List[str] = []
        for r in grp:
            merged_warns.extend(r.get("_warns") or [])
            raw = str(r.get("code") or "").strip()
            if not raw:
                continue
            for fragment in re.split(r"\s*;\s*", raw):
                lbl = fragment.strip()
                if lbl:
                    uniq[lbl] = None
        code_cell = "; ".join(_canonical_sort_code_labels(list(uniq.keys())))
        out.append({"quote": quote, "code": code_cell, "_warns": merged_warns})
    return out


def extract_highlight_rows(path: str) -> List[dict]:
    """
    Preliminary rows: every Word comment anchor, plus highlights that sit *outside*
    any comment anchor (comment balloon optional). Same-quote merge happens later.

    Each dict has: comment_id, quote, code, raw_comment, author, _warns.
    """
    rows_out: List[dict] = []
    with zipfile.ZipFile(path, "r") as zf:
        try:
            doc_xml = zf.read("word/document.xml")
        except KeyError:
            return rows_out
        doc_root = ET.fromstring(doc_xml)
        comments = _parse_comments_xml(zf)
        intervals = _all_comment_anchor_intervals(doc_root, comments.keys())

        def cid_key(cid: str) -> Tuple[int | str, str]:
            return (int(cid) if cid.isdigit() else cid, cid)

        for cid in sorted(comments.keys(), key=cid_key):
            author, comment_body = comments[cid]
            si, ei = _segment_anchor_indices(doc_root, cid)
            ordered_runs: List[ET.Element] = []
            if si is not None and ei is not None and ei > si:
                ordered_runs = _runs_between_anchor_indices(doc_root, si, ei)
            chunks = _merged_highlight_chunks(ordered_runs)
            hl_dom = _dominant_highlight(chunks)
            quote = _quote_for_highlight(chunks, hl_dom) if hl_dom else ""

            raw_seg = _segment_text_for_comment(doc_root, cid)
            raw_seg_norm = _normalize_segment_text(raw_seg)
            quote_norm = quote or _normalize_segment_text(raw_seg)
            quote_norm = _format_transcript_layout(quote_norm)

            code, warns = _code_from_highlight_and_comment(hl_dom, comment_body)

            rows_out.append(
                {
                    "comment_id": cid,
                    "quote": quote_norm,
                    "code": code.strip(),
                    "raw_comment": comment_body,
                    "author": author,
                    "_warns": warns,
                    "_doc_order": si if si is not None else 2**31,
                },
            )

        hl_segments = _segments_highlighted_outside_comments(doc_root, intervals)
        rows_out.extend(_rows_from_highlight_only_segments(hl_segments))

        rows_out.sort(
            key=lambda r: (r["_doc_order"], _cid_sort_tuple(str(r["comment_id"]))),
        )
        for r in rows_out:
            del r["_doc_order"]

    return rows_out


def _write_three_column_sheet(rows: List[dict], output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Coding"
    ws.append(["no", "quote", "code"])
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, row in enumerate(rows, start=1):
        ws.append([i, row["quote"], row["code"]])
        for cell in ws[i + 1]:
            cell.alignment = wrap
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 88
    ws.column_dimensions["C"].width = 40

    out_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def export_highlight_paths_to_xlsx(paths: List[str], output_path: str) -> tuple[int, List[str]]:
    """Write merged no/quote/code export; returns (rows, skipped_basenames)."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise ImportError("Install openpyxl: python -m pip install openpyxl") from e

    all_rows: List[dict] = []
    skipped: List[str] = []
    for p in paths:
        p = os.path.abspath(p)
        if not os.path.isfile(p):
            continue
        label = os.path.basename(p)
        per = _merge_highlight_rows_same_quote(extract_highlight_rows(p))
        if not per:
            skipped.append(label)
            continue
        all_rows.extend(per)

        for r in per:
            for ln in r.pop("_warns", []) or []:
                print(f"{label}: {ln}", file=sys.stderr)

    if not all_rows:
        raise ValueError(
            "No data to export. Add Word comments or body highlights that map to codes.",
        )

    _write_three_column_sheet(all_rows, output_path)
    return len(all_rows), skipped


# -----------------------------------------------------------------------------
# Legacy: one row per anchored segment grouped by normalized text × author columns

def extract_file_legacy(path: str) -> List[Tuple[str, str, str, str]]:
    rows: List[Tuple[str, str, str, str]] = []
    with zipfile.ZipFile(path, "r") as zf:
        try:
            doc_xml = zf.read("word/document.xml")
        except KeyError:
            return rows
        doc_root = ET.fromstring(doc_xml)
        comments = _parse_comments_xml(zf)

        for cid in sorted(comments.keys(), key=lambda x: int(x) if x.isdigit() else x):
            author, code_text = comments[cid]
            raw_seg = _segment_text_for_comment(doc_root, cid)
            seg = _normalize_segment_text(raw_seg)
            rows.append((cid, author, seg, code_text))
    return rows


def _author_key(name: str) -> str:
    n = (name or "").strip()
    return n if n else "(no author)"


def build_segment_rows(
    file_label: str, per_comment: List[Tuple[str, str, str, str]]
) -> List[dict]:
    groups: Dict[str, List[Tuple[str, str, str, str]]] = defaultdict(list)
    for row in per_comment:
        cid, author, seg, code = row
        key = seg if seg else f"__empty__:{cid}"
        groups[key].append(row)

    out: List[dict] = []
    for key in sorted(groups.keys(), key=lambda k: (k.startswith("__empty__"), k)):
        items = groups[key]
        items.sort(key=lambda t: int(t[0]) if t[0].isdigit() else t[0])
        seg_text = items[0][2] if not key.startswith("__empty__:") else ""
        seg_text = _format_transcript_layout(seg_text)

        by_author: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
        for cid, author, _seg, code in items:
            by_author[_author_key(author)].append((cid, code))

        author_codes: Dict[str, str] = {}
        for au, pairs in by_author.items():
            pairs.sort(key=lambda x: int(x[0]) if x[0].isdigit() else x[0])
            merged = "; ".join(c for _, c in pairs if c)
            if merged:
                author_codes[au] = merged

        out.append(
            {
                "File_Name": file_label,
                "Coding_Segment": seg_text,
                "_author_codes": author_codes,
            },
        )
    return out


def _write_coder_matrix_workbook(all_rows: List[dict], output_path: str) -> None:
    author_columns: List[str] = sorted({a for r in all_rows for a in r["_author_codes"].keys()})

    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Coding"
    cols = ["Segment_ID", "File_Name", "Coding_Segment", *author_columns]
    ws.append(cols)
    for r in all_rows:
        ac = r["_author_codes"]
        row_cells = [r["Segment_ID"], r["File_Name"], r["Coding_Segment"]]
        row_cells.extend(ac.get(name, "") for name in author_columns)
        ws.append(row_cells)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    max_col = 3 + len(author_columns)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=max_col):
        for cell in row:
            cell.alignment = wrap_top
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 88
    for i, _name in enumerate(author_columns, start=4):
        ws.column_dimensions[get_column_letter(i)].width = 36

    for r in all_rows:
        del r["_author_codes"]

    out_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def export_coder_matrix_to_xlsx(
    paths: List[str],
    output_path: str,
) -> tuple[int, List[str]]:
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise ImportError("Install openpyxl: python -m pip install openpyxl") from e

    all_rows: List[dict] = []
    skipped: List[str] = []
    seg_id = 1
    for p in paths:
        p = os.path.abspath(p)
        if not os.path.isfile(p):
            continue
        label = os.path.basename(p)
        per = extract_file_legacy(p)
        if not per:
            skipped.append(label)
            continue
        for block in build_segment_rows(label, per):
            block["Segment_ID"] = seg_id
            seg_id += 1
            all_rows.append(block)

    if not all_rows:
        raise ValueError(
            "No data to export. Check that your .docx files contain Word comments.",
        )

    _write_coder_matrix_workbook(all_rows, output_path)
    return len(all_rows), skipped


def export_docx_paths_to_xlsx(
    paths: List[str],
    output_path: str,
    *,
    coder_matrix: bool = False,
) -> tuple[int, List[str]]:
    if coder_matrix:
        return export_coder_matrix_to_xlsx(paths, output_path)
    return export_highlight_paths_to_xlsx(paths, output_path)


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Word coding → Excel. Default: no/quote/code from highlights + comments. "
            "Use --coder-matrix for one column per comment author."
        ),
    )
    ap.add_argument(
        "input",
        nargs="+",
        help="One or more .docx paths (or folders with --dir)",
    )
    ap.add_argument(
        "-d",
        "--dir",
        action="store_true",
        help="Treat inputs as folders; process each *.docx inside",
    )
    ap.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output .xlsx path",
    )
    ap.add_argument(
        "--coder-matrix",
        action="store_true",
        help="Legacy layout: Segment_ID / File_Name / Coding_Segment + one column per author",
    )
    args = ap.parse_args()

    paths: List[str] = []
    if args.dir:
        for d in args.input:
            if not os.path.isdir(d):
                raise SystemExit(f"Not a directory: {d}")
            for name in sorted(os.listdir(d)):
                if name.lower().endswith(".docx") and not name.startswith("~$"):
                    paths.append(os.path.join(d, name))
    else:
        paths = [os.path.abspath(p) for p in args.input]

    if not paths:
        raise SystemExit("No .docx files to process.")

    out_path = os.path.abspath(args.output)
    try:
        n, skipped = export_docx_paths_to_xlsx(
            paths,
            out_path,
            coder_matrix=args.coder_matrix,
        )
    except ValueError as e:
        raise SystemExit(str(e)) from e
    except ImportError as e:
        raise SystemExit(str(e)) from e

    print(f"Wrote {n} row(s) -> {out_path}")
    for s in skipped:
        print(f"[skip] No exportable highlights/comments: {s}")


if __name__ == "__main__":
    main()
